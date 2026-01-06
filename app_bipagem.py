import os
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox
from PIL import Image, ImageTk

from openpyxl import Workbook, load_workbook

import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure

ARQUIVO_XLSX = "controle_bipagem.xlsx"
ABA_BIPADOS = "Bipados"
ABA_ESTOQUE = "Estoque"
TOTAL_GAIOLAS = 50


class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("⚡ Shopee Xpress - Controle de Bipagem")
        self.root.geometry("1400x800")
        
        # Configuração de cores modernas
        self.colors = {
            'primary': '#EE4D2D',      # Laranja Shopee
            'primary_dark': '#D73211',
            'secondary': '#0066CC',     # Azul para seleções
            'success': '#26AA40',       # Verde para sucesso
            'bg_light': '#F5F5F5',      # Fundo claro
            'bg_white': '#FFFFFF',
            'text_dark': '#333333',
            'text_light': '#666666',
            'border': '#E0E0E0'
        }
        
        # Configurar estilo
        self._configurar_estilo()
        self._criar_ui()
        self._carregar_ou_inicializar_excel()
        self._carregar_tabelas_do_excel()
    
    def _configurar_estilo(self):
        """Configura estilos personalizados para widgets ttk"""
        style = ttk.Style()
        style.theme_use('clam')  # Tema moderno
        
        # Estilo para Notebook (abas)
        style.configure('TNotebook', background=self.colors['bg_light'], borderwidth=0)
        style.configure('TNotebook.Tab', 
                       padding=[20, 10],
                       font=('Segoe UI', 10, 'bold'))
        style.map('TNotebook.Tab',
                 background=[('selected', self.colors['primary'])],
                 foreground=[('selected', 'white'), ('!selected', self.colors['text_dark'])])
        
        # Estilo específico para abas de gaiolas (números devem ser visíveis)
        style.configure('Gaiolas.TNotebook', background='white', borderwidth=1)
        style.configure('Gaiolas.TNotebook.Tab', 
                       padding=[15, 8],
                       font=('Segoe UI', 10, 'bold'))
        style.map('Gaiolas.TNotebook.Tab',
                 background=[('selected', 'white'), ('!selected', '#E8E8E8')],
                 foreground=[('selected', 'black'), ('!selected', 'black')])
        
        # Estilo para botões
        style.configure('Primary.TButton',
                       font=('Segoe UI', 10, 'bold'),
                       padding=[15, 8],
                       background=self.colors['primary'],
                       foreground='white')
        style.map('Primary.TButton',
                 background=[('active', self.colors['primary_dark'])])
        
        style.configure('Success.TButton',
                       font=('Segoe UI', 9),
                       padding=[12, 6],
                       background=self.colors['success'])
        
        # Estilo para frames
        style.configure('Card.TFrame', background=self.colors['bg_white'], relief='raised', borderwidth=1)
        style.configure('TFrame', background=self.colors['bg_light'])
        
        # Estilo para labels
        style.configure('Title.TLabel', font=('Segoe UI', 11, 'bold'), foreground=self.colors['text_dark'])
        style.configure('TLabel', font=('Segoe UI', 9), foreground=self.colors['text_light'])

    def _criar_ui(self):
        # Container principal com fundo moderno
        self.root.configure(bg=self.colors['bg_light'])
        
        # Cabeçalho moderno
        header_frame = tk.Frame(self.root, bg=self.colors['primary'], height=70)
        header_frame.pack(fill="x", padx=0, pady=0)
        header_frame.pack_propagate(False)
        
        # Container interno do cabeçalho
        header_content = tk.Frame(header_frame, bg=self.colors['primary'])
        header_content.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Logo e título
        logo_container = tk.Frame(header_content, bg=self.colors['primary'])
        logo_container.pack(side="left")
        
        logo_text = tk.Label(
            logo_container,
            text="⚡ Shopee Xpress",
            font=("Arial", 20, "bold"),
            fg="white",
            bg=self.colors['primary']
        )
        logo_text.pack(anchor="w")
        
        subtitle = tk.Label(
            logo_container,
            text="Sistema de Controle de Bipagem",
            font=("Segoe UI", 11),
            fg="#FFE4DD",
            bg=self.colors['primary']
        )
        subtitle.pack(anchor="w")
        
        # Info de data/hora
        info_label = tk.Label(
            header_content,
            text=datetime.now().strftime("%d/%m/%Y • %H:%M"),
            font=("Segoe UI", 10),
            fg="white",
            bg=self.colors['primary']
        )
        info_label.pack(side="right", pady=10)
        
        # Abas com estilo moderno
        self.nb = ttk.Notebook(self.root)
        self.nb.pack(fill="both", expand=True, padx=15, pady=15)

        # Aba Bipar (principal)
        self.frame_bipar = ttk.Frame(self.nb)
        self.nb.add(self.frame_bipar, text="Bipar")

        # Área de entrada
        topo = ttk.Frame(self.frame_bipar, padding=10)
        topo.pack(fill="x")

        ttk.Label(topo, text="Gaiola:").pack(side="left")
        self.var_gaiola = tk.IntVar(value=1)
        self.cb_gaiola = ttk.Combobox(
            topo,
            width=6,
            textvariable=self.var_gaiola,
            values=list(range(1, TOTAL_GAIOLAS + 1)),
            state="readonly",
        )
        self.cb_gaiola.pack(side="left", padx=(6, 12))

        ttk.Label(topo, text="Código:").pack(side="left")
        self.ent_codigo = ttk.Entry(topo, width=40)
        self.ent_codigo.pack(side="left", padx=(6, 12))
        self.ent_codigo.bind("<Return>", lambda e: self.bipar())

        btn_bipar = ttk.Button(topo, text="Bipar", command=self.bipar)
        btn_bipar.pack(side="left")

        btn_salvar = ttk.Button(topo, text="Salvar", command=self._salvar_excel)
        btn_salvar.pack(side="left", padx=(12, 0))

        # Grid de gaiolas com card moderno
        lista_frame = ttk.Frame(self.frame_bipar, style='Card.TFrame')
        lista_frame.pack(fill="both", expand=True, padx=15, pady=(0, 15))
        
        # Título da seção
        grid_header = ttk.Frame(lista_frame, style='Card.TFrame')
        grid_header.pack(fill="x", padx=20, pady=(15, 10))
        ttk.Label(grid_header, text="📋 Grade de Bipagem", style='Title.TLabel').pack(side="left")
        
        # Container do grid
        grid_container = ttk.Frame(lista_frame, style='Card.TFrame')
        grid_container.pack(fill="both", expand=True, padx=20, pady=(0, 15))
        
        cols_bipar = [f"G{n}" for n in range(1, TOTAL_GAIOLAS + 1)]
        self.tv_bipar = ttk.Treeview(
            grid_container,
            columns=cols_bipar,
            show="tree headings",
            height=20,
            selectmode="none",
        )
        
        # Sistema de seleção de células customizado
        self.celulas_selecionadas = set()  # Armazena (row_id, col_index)
        self.celula_ancora = None  # Âncora para seleção com Shift
        
        # Configura tags com cores modernas
        self.tv_bipar.tag_configure('selected', background=self.colors['secondary'], foreground='white')
        self.tv_bipar.tag_configure('evenrow', background='#FAFAFA')
        self.tv_bipar.tag_configure('oddrow', background='#FFFFFF')
        
        # Configura coluna de índice (#)
        self.tv_bipar.column("#0", width=40, anchor="center", stretch=False)
        self.tv_bipar.heading("#0", text="#")
        
        for n in range(1, TOTAL_GAIOLAS + 1):
            c = f"G{n}"
            self.tv_bipar.heading(c, text=str(n))
            self.tv_bipar.column(c, width=100, anchor="center", stretch=False)
        
        # Eventos para seleção de células individuais
        self.tv_bipar.bind("<Button-1>", self._selecionar_celula)
        self.tv_bipar.bind("<Control-Button-1>", self._selecionar_celula_ctrl)
        self.tv_bipar.bind("<Control-a>", self._selecionar_todas_celulas)
        self.tv_bipar.bind("<Control-A>", self._selecionar_todas_celulas)
        self.tv_bipar.bind("<Delete>", lambda e: self.apagar_bipagem())
        self.tv_bipar.bind("<Button-3>", self._menu_contexto_bipar)
        self.tv_bipar.bind("<Escape>", lambda e: self._limpar_selecao())
        
        # Eventos Shift+Arrows para expandir seleção
        self.tv_bipar.bind("<Shift-Down>", self._expandir_selecao_baixo)
        self.tv_bipar.bind("<Shift-Up>", self._expandir_selecao_cima)
        self.tv_bipar.bind("<Shift-Left>", self._expandir_selecao_esquerda)
        self.tv_bipar.bind("<Shift-Right>", self._expandir_selecao_direita)

        sbx = ttk.Scrollbar(grid_container, orient="horizontal", command=self.tv_bipar.xview)
        sby = ttk.Scrollbar(grid_container, orient="vertical", command=self.tv_bipar.yview)
        self.tv_bipar.configure(xscrollcommand=sbx.set, yscrollcommand=sby.set)

        self.tv_bipar.grid(row=0, column=0, sticky="nsew")
        sby.grid(row=0, column=1, sticky="ns")
        sbx.grid(row=1, column=0, sticky="ew")

        grid_container.rowconfigure(0, weight=1)
        grid_container.columnconfigure(0, weight=1)

        # Botões de controle modernos
        btn_frame = ttk.Frame(lista_frame, style='Card.TFrame')
        btn_frame.pack(fill="x", padx=20, pady=(0, 15))
        
        btn_apagar = ttk.Button(btn_frame, text="🗑️ Apagar Selecionadas [Del]", 
                               command=self.apagar_bipagem, style='Primary.TButton')
        btn_apagar.pack(side="left", padx=(0, 10))
        
        ttk.Label(btn_frame, 
                 text="💡 Dicas: Clique nas células | Shift+Setas = múltiplas | Ctrl+A = todas | Esc = limpar", 
                 foreground=self.colors['text_light'],
                 font=('Segoe UI', 9)).pack(side="left", padx=10)

        btn_salvar_grid = ttk.Button(btn_frame, text="💾 Salvar", 
                                     command=self._salvar_excel, style='Success.TButton')
        btn_salvar_grid.pack(side="right")

        # Aba Bipados
        self.frame_bipados = ttk.Frame(self.nb, style='TFrame')
        self.nb.add(self.frame_bipados, text="🔲  Bipados (Gaiolas)")

        cols_bipados = [f"G{n}" for n in range(1, TOTAL_GAIOLAS + 1)]
        self.tv_bipados = ttk.Treeview(
            self.frame_bipados,
            columns=cols_bipados,
            show="headings",
            height=18,
        )
        for n in range(1, TOTAL_GAIOLAS + 1):
            c = f"G{n}"
            self.tv_bipados.heading(c, text=str(n))
            self.tv_bipados.column(c, width=80, anchor="center", stretch=False)

        sbx = ttk.Scrollbar(self.frame_bipados, orient="horizontal", command=self.tv_bipados.xview)
        sby = ttk.Scrollbar(self.frame_bipados, orient="vertical", command=self.tv_bipados.yview)
        self.tv_bipados.configure(xscrollcommand=sbx.set, yscrollcommand=sby.set)

        self.tv_bipados.grid(row=0, column=0, sticky="nsew")
        sby.grid(row=0, column=1, sticky="ns")
        sbx.grid(row=1, column=0, sticky="ew")

        self.frame_bipados.rowconfigure(0, weight=1)
        self.frame_bipados.columnconfigure(0, weight=1)

        # Aba Estoque (log)
        self.frame_estoque = ttk.Frame(self.nb, style='TFrame')
        self.nb.add(self.frame_estoque, text="📊  Estoque (Log)")

        cols_estoque = ["data_hora", "codigo", "gaiola"]
        self.tv_estoque = ttk.Treeview(
            self.frame_estoque,
            columns=cols_estoque,
            show="headings",
            height=18,
        )
        self.tv_estoque.heading("data_hora", text="Data/Hora")
        self.tv_estoque.heading("codigo", text="Código")
        self.tv_estoque.heading("gaiola", text="Gaiola")

        self.tv_estoque.column("data_hora", width=180, anchor="w")
        self.tv_estoque.column("codigo", width=240, anchor="w")
        self.tv_estoque.column("gaiola", width=80, anchor="center")

        sby2 = ttk.Scrollbar(self.frame_estoque, orient="vertical", command=self.tv_estoque.yview)
        self.tv_estoque.configure(yscrollcommand=sby2.set)

        self.tv_estoque.grid(row=0, column=0, sticky="nsew")
        sby2.grid(row=0, column=1, sticky="ns")

        self.frame_estoque.rowconfigure(0, weight=1)
        self.frame_estoque.columnconfigure(0, weight=1)

        # Aba Estoque por Gaiola
        self.frame_estoque_gaiolas = ttk.Frame(self.nb, style='TFrame')
        self.nb.add(self.frame_estoque_gaiolas, text="📦  Estoque por Gaiola")
        
        # Criar notebook interno para as gaiolas
        self.nb_gaiolas = ttk.Notebook(self.frame_estoque_gaiolas, style='Gaiolas.TNotebook')
        self.nb_gaiolas.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Criar uma aba para cada gaiola
        self.tv_gaiolas = {}  # Dicionário para armazenar as treeviews de cada gaiola
        
        for g in range(1, TOTAL_GAIOLAS + 1):
            frame_g = ttk.Frame(self.nb_gaiolas)
            self.nb_gaiolas.add(frame_g, text=f" {g} ")  # Número da gaiola
            
            cols_gaiola = ["data_hora", "codigo"]
            tv_g = ttk.Treeview(
                frame_g,
                columns=cols_gaiola,
                show="headings",
                height=18,
            )
            tv_g.heading("data_hora", text="Data/Hora")
            tv_g.heading("codigo", text="Código")
            tv_g.column("data_hora", width=180, anchor="w")
            tv_g.column("codigo", width=400, anchor="w")
            
            sby_g = ttk.Scrollbar(frame_g, orient="vertical", command=tv_g.yview)
            tv_g.configure(yscrollcommand=sby_g.set)
            
            tv_g.grid(row=0, column=0, sticky="nsew")
            sby_g.grid(row=0, column=1, sticky="ns")
            
            frame_g.rowconfigure(0, weight=1)
            frame_g.columnconfigure(0, weight=1)
            
            self.tv_gaiolas[g] = tv_g

        # Aba Estatísticas/Gráficos
        self.frame_stats = ttk.Frame(self.nb, style='TFrame')
        self.nb.add(self.frame_stats, text="📊  Estatísticas")
        
        # Frame para controles
        controles_frame = ttk.Frame(self.frame_stats, padding=10)
        controles_frame.pack(fill="x")
        
        btn_atualizar_grafico = ttk.Button(controles_frame, text="Atualizar Gráficos", command=self.atualizar_graficos)
        btn_atualizar_grafico.pack(side="left")
        
        # Frame para os gráficos
        self.graficos_frame = ttk.Frame(self.frame_stats)
        self.graficos_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # Barra de status
        self.status = tk.StringVar(value="Pronto.")
        ttk.Label(self.root, textvariable=self.status, padding=(10, 0, 10, 10)).pack(fill="x")

        self.root.after(100, self.ent_codigo.focus_set)
        
        # Selecionar a aba Bipar ao iniciar
        self.nb.select(0)

    # -------- Excel --------
    def _carregar_ou_inicializar_excel(self):
        if os.path.exists(ARQUIVO_XLSX):
            self.wb = load_workbook(ARQUIVO_XLSX)
        else:
            self.wb = Workbook()
            # Remove sheet padrão
            if "Sheet" in self.wb.sheetnames:
                del self.wb["Sheet"]

            ws_b = self.wb.create_sheet(ABA_BIPADOS)
            ws_e = self.wb.create_sheet(ABA_ESTOQUE)

            # Cabeçalhos Bipados: 1..50
            for col in range(1, TOTAL_GAIOLAS + 1):
                ws_b.cell(row=1, column=col, value=str(col))

            # Cabeçalhos Estoque
            ws_e.append(["data_hora", "codigo", "gaiola"])

            self.wb.save(ARQUIVO_XLSX)

        # Garante abas
        if ABA_BIPADOS not in self.wb.sheetnames:
            ws_b = self.wb.create_sheet(ABA_BIPADOS)
            for col in range(1, TOTAL_GAIOLAS + 1):
                ws_b.cell(row=1, column=col, value=str(col))

        if ABA_ESTOQUE not in self.wb.sheetnames:
            ws_e = self.wb.create_sheet(ABA_ESTOQUE)
            ws_e.append(["data_hora", "codigo", "gaiola"])

    def _salvar_excel(self):
        self.wb.save(ARQUIVO_XLSX)
        self.status.set(f"Salvo em {ARQUIVO_XLSX} às {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.")

    def _carregar_tabelas_do_excel(self):
        # Limpa treeviews
        for tv in (self.tv_bipados, self.tv_estoque, self.tv_bipar):
            for iid in tv.get_children():
                tv.delete(iid)
        
        # Limpa seleções de células
        self.celulas_selecionadas.clear()
        
        # Limpa treeviews das gaiolas
        for tv_g in self.tv_gaiolas.values():
            for iid in tv_g.get_children():
                tv_g.delete(iid)

        ws_b = self.wb[ABA_BIPADOS]
        ws_e = self.wb[ABA_ESTOQUE]

        # Carrega Bipados (linhas 2..max)
        max_row = ws_b.max_row
        # Monta lista de linhas com 50 colunas
        linhas = []
        for r in range(2, max_row + 1):
            vals = []
            for c in range(1, TOTAL_GAIOLAS + 1):
                v = ws_b.cell(row=r, column=c).value
                vals.append("" if v is None else str(v))
            # Só adiciona se tiver algum valor
            if any(v != "" for v in vals):
                linhas.append(vals)

        # Mostra linhas
        for row in linhas:
            self.tv_bipados.insert("", "end", values=row)

        # Carrega Estoque e aba Bipar (linhas 2..)
        todos_registros = []
        for r in range(2, ws_e.max_row + 1):
            data_hora = ws_e.cell(row=r, column=1).value
            codigo = ws_e.cell(row=r, column=2).value
            gaiola = ws_e.cell(row=r, column=3).value
            if data_hora is None and codigo is None and gaiola is None:
                continue
            registro = [
                "" if data_hora is None else str(data_hora),
                "" if codigo is None else str(codigo),
                "" if gaiola is None else str(gaiola),
            ]
            todos_registros.append((r, registro))  # Guarda também a linha do Excel
            self.tv_estoque.insert("", "end", values=registro)
            
            # Adiciona na treeview da gaiola correspondente
            try:
                gaiola_num = int(registro[2]) if registro[2] else 0
                if 1 <= gaiola_num <= TOTAL_GAIOLAS:
                    self.tv_gaiolas[gaiola_num].insert("", "end", values=[registro[0], registro[1]])
            except (ValueError, KeyError):
                pass
        
        # Carrega aba Bipar (igual à aba Bipados) com numeração de linhas
        linha_numero = 1
        for row in linhas:
            self.tv_bipar.insert("", "end", text=str(linha_numero), values=row)
            linha_numero += 1

    # -------- Regras de bipagem --------
    def _proxima_linha_vazia_na_coluna(self, ws, col_index: int) -> int:
        """
        Retorna a próxima linha (>=2) vazia para a coluna col_index.
        """
        r = 2
        while True:
            if ws.cell(row=r, column=col_index).value in (None, ""):
                return r
            r += 1

    def bipar(self):
        codigo = self.ent_codigo.get().strip()
        if not codigo:
            return

        try:
            gaiola = int(self.var_gaiola.get())
        except Exception:
            messagebox.showerror("Erro", "Selecione uma gaiola válida (1..50).")
            return

        if not (1 <= gaiola <= TOTAL_GAIOLAS):
            messagebox.showerror("Erro", "Gaiola inválida (1..50).")
            return

        ws_b = self.wb[ABA_BIPADOS]
        ws_e = self.wb[ABA_ESTOQUE]

        col = gaiola  # 1..50
        linha = self._proxima_linha_vazia_na_coluna(ws_b, col)
        ws_b.cell(row=linha, column=col, value=codigo)

        data_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_e.append([data_hora, codigo, gaiola])

        # Atualiza UI (sem precisar recarregar tudo)
        # Garante que a Treeview de bipados tenha linhas suficientes
        current_rows = len(self.tv_bipados.get_children())
        target_row_index = linha - 2  # 0-based dentro da treeview
        while current_rows <= target_row_index:
            self.tv_bipados.insert("", "end", values=[""] * TOTAL_GAIOLAS)
            current_rows += 1

        iid = self.tv_bipados.get_children()[target_row_index]
        row_values = list(self.tv_bipados.item(iid, "values"))
        row_values[col - 1] = codigo
        self.tv_bipados.item(iid, values=row_values)

        self.tv_estoque.insert("", "end", values=[data_hora, codigo, str(gaiola)])
        self.tv_estoque.yview_moveto(1.0)
        
        # Atualiza aba Bipar (mesmo formato que Bipados)
        current_rows_bipar = len(self.tv_bipar.get_children())
        while current_rows_bipar <= target_row_index:
            linha_numero = current_rows_bipar + 1
            self.tv_bipar.insert("", "end", text=str(linha_numero), values=[""] * TOTAL_GAIOLAS)
            current_rows_bipar += 1

        iid_bipar = self.tv_bipar.get_children()[target_row_index]
        row_values_bipar = list(self.tv_bipar.item(iid_bipar, "values"))
        row_values_bipar[col - 1] = codigo
        self.tv_bipar.item(iid_bipar, values=row_values_bipar)
        
        # Adiciona na treeview da gaiola específica
        if 1 <= gaiola <= TOTAL_GAIOLAS:
            self.tv_gaiolas[gaiola].insert("", "end", values=[data_hora, codigo])

        self.ent_codigo.delete(0, "end")
        self.ent_codigo.focus_set()

        # Salva automaticamente
        self._salvar_excel()

    def apagar_bipagem(self):
        """Apaga as células selecionadas da aba Bipar"""
        if not self.celulas_selecionadas:
            messagebox.showwarning("Aviso", "Clique nas células que deseja apagar.")
            return
        
        ws_b = self.wb[ABA_BIPADOS]
        ws_e = self.wb[ABA_ESTOQUE]
        
        # Coleta todas as células selecionadas com conteúdo
        itens_para_apagar = []
        
        for row_id, col_idx in self.celulas_selecionadas:
            valores = self.tv_bipar.item(row_id, "values")
            if col_idx < len(valores):
                valor = valores[col_idx]
                if valor and str(valor).strip():
                    gaiola = col_idx + 1  # Gaiolas começam em 1
                    itens_para_apagar.append({
                        "codigo": str(valor),
                        "gaiola": gaiola,
                        "row_id": row_id,
                        "col_idx": col_idx
                    })
        
        if not itens_para_apagar:
            messagebox.showwarning("Aviso", "Nenhuma célula com conteúdo foi selecionada.")
            return
        
        # Confirma exclusão
        qtd = len(itens_para_apagar)
        if qtd == 1:
            item = itens_para_apagar[0]
            msg = f"Apagar item?\n\nCódigo: {item['codigo']}\nGaiola: {item['gaiola']}"
        else:
            msg = f"Apagar {qtd} itens selecionados?\n\nEsta ação não pode ser desfeita."
        
        resposta = messagebox.askyesno("Confirmar exclusão", msg)
        if not resposta:
            return
        
        # Remove cada item
        for item in itens_para_apagar:
            codigo = item['codigo']
            gaiola = item['gaiola']
            
            # Remove da aba de bipados - procura o código na coluna da gaiola
            col = gaiola
            for r in range(2, ws_b.max_row + 1):
                valor_celula = ws_b.cell(row=r, column=col).value
                if str(valor_celula) == codigo:
                    ws_b.cell(row=r, column=col, value="")
                    break
            
            # Remove do log de estoque - encontra e remove a linha correspondente
            for r in range(ws_e.max_row, 1, -1):  # Começa do fim
                cod_estoque = ws_e.cell(row=r, column=2).value
                gaiola_estoque = ws_e.cell(row=r, column=3).value
                
                if str(cod_estoque) == codigo and str(gaiola_estoque) == str(gaiola):
                    ws_e.delete_rows(r, 1)
                    break
        
        # Salva e recarrega
        self._salvar_excel()
        self._carregar_tabelas_do_excel()
        
        if qtd == 1:
            self.status.set(f"Item apagado: {itens_para_apagar[0]['codigo']} (Gaiola {itens_para_apagar[0]['gaiola']})")
        else:
            self.status.set(f"{qtd} itens apagados com sucesso")
        
        self.ent_codigo.focus_set()
    
    def _selecionar_celula(self, event):
        """Seleciona uma célula individual (substitui seleção anterior)"""
        region = self.tv_bipar.identify_region(event.x, event.y)
        if region != "cell":
            return
        
        row_id = self.tv_bipar.identify_row(event.y)
        col_id = self.tv_bipar.identify_column(event.x)
        
        if not row_id or not col_id:
            return
        
        # Verifica se a célula tem conteúdo
        col_idx = int(col_id.replace("#", "")) - 1
        valores = self.tv_bipar.item(row_id, "values")
        
        if col_idx >= len(valores) or not valores[col_idx] or not str(valores[col_idx]).strip():
            return  # Não seleciona células vazias
        
        # Limpa seleção anterior
        self._limpar_selecao()
        
        # Adiciona nova seleção
        self.celulas_selecionadas.add((row_id, col_idx))
        self.celula_ancora = (row_id, col_idx)  # Define âncora para Shift+Arrow
        self._atualizar_visual_selecao()
    
    def _selecionar_celula_ctrl(self, event):
        """Seleciona célula adicional com Ctrl (seleção múltipla)"""
        region = self.tv_bipar.identify_region(event.x, event.y)
        if region != "cell":
            return
        
        row_id = self.tv_bipar.identify_row(event.y)
        col_id = self.tv_bipar.identify_column(event.x)
        
        if not row_id or not col_id:
            return
        
        # Verifica se a célula tem conteúdo
        col_idx = int(col_id.replace("#", "")) - 1
        valores = self.tv_bipar.item(row_id, "values")
        
        if col_idx >= len(valores) or not valores[col_idx] or not str(valores[col_idx]).strip():
            return  # Não seleciona células vazias
        
        # Toggle da célula
        celula = (row_id, col_idx)
        if celula in self.celulas_selecionadas:
            self.celulas_selecionadas.remove(celula)
            if celula == self.celula_ancora:
                self.celula_ancora = None
        else:
            self.celulas_selecionadas.add(celula)
            self.celula_ancora = (row_id, col_idx)  # Atualiza âncora
        
        self._atualizar_visual_selecao()
    
    def _selecionar_todas_celulas(self, event=None):
        """Seleciona todas as células com conteúdo"""
        self.celulas_selecionadas.clear()
        
        for row_id in self.tv_bipar.get_children():
            valores = self.tv_bipar.item(row_id, "values")
            for col_idx, valor in enumerate(valores):
                if valor and str(valor).strip():
                    self.celulas_selecionadas.add((row_id, col_idx))
        
        self._atualizar_visual_selecao()
        return "break"
    
    def _limpar_selecao(self, event=None):
        """Limpa todas as seleções de células"""
        self.celulas_selecionadas.clear()
        self.celula_ancora = None
        self._atualizar_visual_selecao()
    
    def _atualizar_visual_selecao(self):
        """Atualiza a visualização das linhas selecionadas"""
        # Remove tag 'selected' de todas as linhas
        for item in self.tv_bipar.get_children():
            tags = list(self.tv_bipar.item(item, 'tags'))
            if 'selected' in tags:
                tags.remove('selected')
            self.tv_bipar.item(item, tags=tags)
        
        # Aplica tag 'selected' nas linhas que contêm células selecionadas
        rows_selecionadas = set()
        for row_id, col_idx in self.celulas_selecionadas:
            rows_selecionadas.add(row_id)
        
        for row_id in rows_selecionadas:
            try:
                tags = list(self.tv_bipar.item(row_id, 'tags'))
                if 'selected' not in tags:
                    tags.append('selected')
                self.tv_bipar.item(row_id, tags=tags)
            except:
                pass
    
    def _expandir_selecao_baixo(self, event):
        """Expande seleção para baixo com Shift+Down"""
        if not self.celula_ancora:
            return "break"
        
        row_ancora, col_ancora = self.celula_ancora
        todas_linhas = self.tv_bipar.get_children()
        
        try:
            idx_ancora = todas_linhas.index(row_ancora)
            # Próxima linha
            if idx_ancora < len(todas_linhas) - 1:
                proxima_linha = todas_linhas[idx_ancora + 1]
                valores = self.tv_bipar.item(proxima_linha, "values")
                if col_ancora < len(valores) and valores[col_ancora] and str(valores[col_ancora]).strip():
                    self.celulas_selecionadas.add((proxima_linha, col_ancora))
                    self.celula_ancora = (proxima_linha, col_ancora)
                    self._atualizar_visual_selecao()
        except:
            pass
        return "break"
    
    def _expandir_selecao_cima(self, event):
        """Expande seleção para cima com Shift+Up"""
        if not self.celula_ancora:
            return "break"
        
        row_ancora, col_ancora = self.celula_ancora
        todas_linhas = self.tv_bipar.get_children()
        
        try:
            idx_ancora = todas_linhas.index(row_ancora)
            # Linha anterior
            if idx_ancora > 0:
                linha_anterior = todas_linhas[idx_ancora - 1]
                valores = self.tv_bipar.item(linha_anterior, "values")
                if col_ancora < len(valores) and valores[col_ancora] and str(valores[col_ancora]).strip():
                    self.celulas_selecionadas.add((linha_anterior, col_ancora))
                    self.celula_ancora = (linha_anterior, col_ancora)
                    self._atualizar_visual_selecao()
        except:
            pass
        return "break"
    
    def _expandir_selecao_direita(self, event):
        """Expande seleção para direita com Shift+Right"""
        if not self.celula_ancora:
            return "break"
        
        row_ancora, col_ancora = self.celula_ancora
        
        try:
            valores = self.tv_bipar.item(row_ancora, "values")
            # Próxima coluna
            if col_ancora < len(valores) - 1:
                nova_col = col_ancora + 1
                if valores[nova_col] and str(valores[nova_col]).strip():
                    self.celulas_selecionadas.add((row_ancora, nova_col))
                    self.celula_ancora = (row_ancora, nova_col)
                    self._atualizar_visual_selecao()
        except:
            pass
        return "break"
    
    def _expandir_selecao_esquerda(self, event):
        """Expande seleção para esquerda com Shift+Left"""
        if not self.celula_ancora:
            return "break"
        
        row_ancora, col_ancora = self.celula_ancora
        
        try:
            valores = self.tv_bipar.item(row_ancora, "values")
            # Coluna anterior
            if col_ancora > 0:
                nova_col = col_ancora - 1
                if valores[nova_col] and str(valores[nova_col]).strip():
                    self.celulas_selecionadas.add((row_ancora, nova_col))
                    self.celula_ancora = (row_ancora, nova_col)
                    self._atualizar_visual_selecao()
        except:
            pass
        return "break"
    
    def _menu_contexto_bipar(self, event):
        """Menu de contexto com botão direito"""
        menu = tk.Menu(self.root, tearoff=0)
        menu.add_command(label="Apagar selecionada(s)", command=self.apagar_bipagem)
        menu.add_command(label="Selecionar todas (Ctrl+A)", command=self._selecionar_todas_celulas)
        menu.add_command(label="Limpar seleção (Esc)", command=self._limpar_selecao)
        menu.add_separator()
        menu.add_command(label="Cancelar")
        
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()
    
    def atualizar_graficos(self):
        """Atualiza os gráficos de estatísticas"""
        # Limpa gráficos anteriores
        for widget in self.graficos_frame.winfo_children():
            widget.destroy()
        
        ws_e = self.wb[ABA_ESTOQUE]
        
        # Conta itens por gaiola
        contagem_gaiolas = {g: 0 for g in range(1, TOTAL_GAIOLAS + 1)}
        total_itens = 0
        
        for r in range(2, ws_e.max_row + 1):
            gaiola = ws_e.cell(row=r, column=3).value
            if gaiola is not None:
                try:
                    g_num = int(gaiola)
                    if 1 <= g_num <= TOTAL_GAIOLAS:
                        contagem_gaiolas[g_num] += 1
                        total_itens += 1
                except ValueError:
                    pass
        
        # Cria figura com múltiplos gráficos
        fig = Figure(figsize=(12, 8))
        
        # Gráfico 1: Barras - Itens por Gaiola
        ax1 = fig.add_subplot(2, 2, 1)
        gaiolas_com_itens = {g: count for g, count in contagem_gaiolas.items() if count > 0}
        if gaiolas_com_itens:
            ax1.bar(gaiolas_com_itens.keys(), gaiolas_com_itens.values(), color='steelblue')
            ax1.set_xlabel('Gaiola')
            ax1.set_ylabel('Quantidade de Itens')
            ax1.set_title('Distribuição de Itens por Gaiola')
            ax1.grid(axis='y', alpha=0.3)
        else:
            ax1.text(0.5, 0.5, 'Sem dados', ha='center', va='center', transform=ax1.transAxes)
            ax1.set_title('Distribuição de Itens por Gaiola')
        
        # Gráfico 2: Pizza - Top 10 Gaiolas
        ax2 = fig.add_subplot(2, 2, 2)
        top_10 = sorted(gaiolas_com_itens.items(), key=lambda x: x[1], reverse=True)[:10]
        if top_10:
            labels = [f"Gaiola {g}" for g, _ in top_10]
            sizes = [count for _, count in top_10]
            ax2.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90)
            ax2.set_title('Top 10 Gaiolas com Mais Itens')
        else:
            ax2.text(0.5, 0.5, 'Sem dados', ha='center', va='center', transform=ax2.transAxes)
            ax2.set_title('Top 10 Gaiolas com Mais Itens')
        
        # Gráfico 3: Resumo textual
        ax3 = fig.add_subplot(2, 2, 3)
        ax3.axis('off')
        
        gaiolas_ocupadas = len(gaiolas_com_itens)
        gaiolas_vazias = TOTAL_GAIOLAS - gaiolas_ocupadas
        media_itens = total_itens / gaiolas_ocupadas if gaiolas_ocupadas > 0 else 0
        
        resumo_texto = f"""
        RESUMO GERAL
        
        Total de Itens: {total_itens}
        Gaiolas Ocupadas: {gaiolas_ocupadas}
        Gaiolas Vazias: {gaiolas_vazias}
        Média por Gaiola: {media_itens:.2f}
        
        Gaiola Mais Cheia: {top_10[0][0] if top_10 else 'N/A'} ({top_10[0][1] if top_10 else 0} itens)
        """
        
        ax3.text(0.1, 0.5, resumo_texto, fontsize=12, verticalalignment='center',
                family='monospace', bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.5))
        
        # Gráfico 4: Linha - Gaiolas vazias vs ocupadas
        ax4 = fig.add_subplot(2, 2, 4)
        categorias = ['Ocupadas', 'Vazias']
        valores = [gaiolas_ocupadas, gaiolas_vazias]
        cores = ['green', 'red']
        ax4.bar(categorias, valores, color=cores, alpha=0.7)
        ax4.set_ylabel('Quantidade')
        ax4.set_title('Status das Gaiolas')
        ax4.grid(axis='y', alpha=0.3)
        
        for i, v in enumerate(valores):
            ax4.text(i, v + 1, str(v), ha='center', va='bottom', fontweight='bold')
        
        fig.tight_layout()
        
        # Adiciona o canvas ao frame
        canvas = FigureCanvasTkAgg(fig, master=self.graficos_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)
        
        self.status.set(f"Gráficos atualizados - Total: {total_itens} itens em {gaiolas_ocupadas} gaiolas")


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()